"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
import platform

# File paths
input_file = 'OfferReport.xlsx'  # Replace with your input file
output_file = 'NewHire_IT_Checklist.xlsx'  # Replace with your output file

# Read and process the Excel file from the sheet named "FTE Hire"
df = pd.read_excel(input_file, sheet_name='FTE Hire')

# Strip any extra spaces from column names
df.columns = df.columns.str.strip()

# Function to generate email alias
def generate_email_alias(full_name):
    names = full_name.split()
    if len(names) < 2:
        return ""  # Handle cases where there's no last name
    first_name = names[0].lower()
    last_name_initial = names[-1][0].lower()  # Get the first letter of the last name
    email_alias = f"{first_name}{last_name_initial}@better.com"
    return email_alias

# Function to extract username (remove "@better.com")
def extract_username(work_email):
    if pd.isna(work_email) or not isinstance(work_email, str):
        return ""  # Return empty string if the email is missing or not a string
    return work_email.split('@')[0]

# Main function to create the output Excel sheet
def create_excel(start_row, end_row, input_df, output_file):
     # Adjust the rows to be zero-based for pandas indexing
    selected_data = input_df.iloc[start_row - 1:end_row]

    # Correct the column names based on the actual file's structure
    columns_needed = ['Candidate', 'Personal Email', 'Better Email', 'Job', 'Department', 'Office', 'Start Date', 'Office Location (for OL)', 'TimeZone']
    selected_data = selected_data[columns_needed]


    # Create a new DataFrame for the output Excel sheet
    output_data = pd.DataFrame({
        'Full Name': selected_data['Candidate'],
        'Personal Email': selected_data['Personal Email'],
        'Username': selected_data['Better Email'].apply(extract_username),  # Remove "@better.com"
        'Better Email': selected_data['Better Email'],
        'Temporary Password': '',  # Leave this column empty
        'Title': selected_data['Job'],
        'Department': selected_data['Department'],
        'Front Setup Needed': selected_data['Department'].apply(lambda dept: 'Yes' if dept == 'SPOC' else 'N/A'),
        'Front Inbox Setup Level': 'Personal',  # Print "Personal" for each row
        'Email Alias Paragraph': selected_data['Department'].apply(lambda dept: 'This is your External Email Alias, please do not refer to this Alias email address until we meet in the IT Onboarding Session:' if dept == 'SPOC' else ''),
        'Email Alias': selected_data.apply(lambda row: generate_email_alias(row['Candidate']) if row['Department'] == 'SPOC' else 'N/A', axis=1),  # Generate email alias for SPOC
        'Location': selected_data['Office Location (for OL)'],
        'Start Date': pd.to_datetime(selected_data['Start Date'], errors='coerce').dt.strftime('%B %d, %Y'),  # Change date format to (Month, Day, Year)
        'Start Time': selected_data['TimeZone'].apply(lambda tz: '10:30 AM EST' if 'EST' in str(tz) else '12:30 AM PST' if 'PST' in str(tz) else ''),  # Set time based on TimeZone
        'Zoom Link': '',  
        'Zoom ID/Pass': '',  
        'Monitor FedEx Tracking': '',
        'WFH Bundle FedEx Tracking': '', 
        'Laptop FedEx Tracking': ''  
    })

    # Generate a sheet name based on the start date
    first_start_date = pd.to_datetime(selected_data['Start Date'].iloc[0], errors='coerce')

    # Check if the first_start_date is valid, otherwise set a default sheet name
    if pd.notna(first_start_date):
        sheet_name = f'Sheet_{first_start_date.strftime("%B_%d_%Y")}'  # Use (Month, Day, Year) format for sheet name
    else:
        sheet_name = 'Sheet_NoStartDate'

    # Load the existing workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(output_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Check if the sheet already exists, and if so, remove it
    if sheet_name in workbook.sheetnames:
        std = workbook[sheet_name]
        workbook.remove(std)

    # Add a new sheet to the workbook
    sheet = workbook.create_sheet(title=sheet_name)

    # Write DataFrame to the new sheet
    for r_idx, row in enumerate(dataframe_to_rows(output_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            # Apply border around each cell
            cell.border = Border(left=Side(border_style='thin'),
                                 right=Side(border_style='thin'),
                                 top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'))

    # Define the styles
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light blue color
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Apply styles to the header row
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Adjust column widths, row heights, and center align all cells
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = center_alignment
            cell.border = Border(left=Side(border_style='thin'),
                                 right=Side(border_style='thin'),
                                 top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'))
        sheet.column_dimensions[column_letter].width = max_length + 5  # Add padding for better readability

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
            cell.border = Border(left=Side(border_style='thin'),
                                 right=Side(border_style='thin'),
                                 top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'))
        sheet.row_dimensions[row[0].row].height = 30  # Set row height

    # Hide columns that are empty
    for col in sheet.columns:
        if all(cell.value is None for cell in col):
            sheet.column_dimensions[col[0].column_letter].hidden = True

    # Save the workbook with the new sheet and formatting
    workbook.save(output_file)

    print(f"New Excel sheet '{sheet_name}' created and formatted successfully in '{output_file}'!")

    # Open the output file
    open_excel_file(output_file)

# Define the open function based on the operating system
def open_excel_file(file_path):
    if platform.system() == 'Darwin':  # macOS
        subprocess.call(['open', file_path])
    elif platform.system() == 'Windows':  # Windows
        import os
        os.startfile(file_path)
    else:  # Linux
        subprocess.call(['xdg-open', file_path])

# Example usage:
start_row = 327  # Specify the start row (zero-based index)
end_row = 347  # Specify the end row (zero-based index)

create_excel(start_row, end_row, df, output_file)
"""